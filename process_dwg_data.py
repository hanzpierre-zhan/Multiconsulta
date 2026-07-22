import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_layer_name(name):
    if not isinstance(name, str):
        return name
    # Clean encoding issues common in CAD layer names
    replacements = {
        'FO': 'F.O.',
        'F\ufffdO\ufffd': 'F.O.',
        'COTAS PY': 'COTAS PROYECTO',
    }
    for old, new in replacements.items():
        name = name.replace(old, new)
    return name

def main():
    csv_path = r"C:\Users\PLANTA EXTERNA\Desktop\mygeodata_extracted\ULTIMO DE JULIO.csv"
    excel_dir = r"C:\Users\PLANTA EXTERNA\Desktop\Nueva carpeta"
    excel_path = os.path.join(excel_dir, "informacion_autocad.xlsx")

    print("Cargando archivo CSV...")
    df = pd.read_csv(csv_path)

    # Clean layer names
    df['Layer'] = df['Layer'].apply(clean_layer_name)

    # Ensure coordinates are numeric
    df['X'] = pd.to_numeric(df['X'], errors='coerce')
    df['Y'] = pd.to_numeric(df['Y'], errors='coerce')
    df['Z'] = pd.to_numeric(df['Z'], errors='coerce')

    print("Procesando datos por categorías...")

    # Category 1: Summary of Layers
    summary_df = df['Layer'].value_counts().reset_index()
    summary_df.columns = ['Capa (Layer)', 'Cantidad de Elementos']
    
    # Add descriptions to layers
    descriptions = {
        '0': 'Capa base / por defecto',
        'Unknown_Point_Feature': 'Puntos de características de terreno',
        'FEATURE_LABEL': 'Etiquetas e identificación de elementos importantes (ej. puentes, pueblos)',
        'POSTES PROTECTADOS': 'Postes con protección física',
        'POSTES ELECTRICOS': 'Postes de la red eléctrica comercial',
        'F.O. COTAS': 'Cotas y distancias de tendido de Fibra Óptica',
        'DEFPOINTS': 'Puntos de definición del dibujo y etiquetas de estaciones',
        'KILOMETRAJE': 'Marcadores de kilómetro de la vía',
        'F.O. RETENIDAS': 'Retenidas o tensores de fibra óptica',
        'COTAS PROYECTO': 'Cotas y elevaciones del proyecto',
        'F.O. CABLE': 'Trazados de cables de fibra óptica',
        'EJE DE VIA': 'Línea de centro de la carretera/vía',
        'CARRETERA': 'Bordes y trazado de la carretera',
        'DERECHO DE VIA': 'Límites de faja de servidumbre / derecho de vía'
    }
    summary_df['Descripción'] = summary_df['Capa (Layer)'].map(descriptions).fillna('Otros elementos del plano')

    # Category 2: Poles and Points
    poles_layers = ['POSTES ELECTRICOS', 'POSTES PROTECTADOS', 'Unknown_Point_Feature']
    poles_df = df[df['Layer'].isin(poles_layers)].copy()
    poles_df = poles_df[['X', 'Y', 'Z', 'Layer', 'SubClasses', 'EntityHandle']].sort_values(by=['Layer', 'X'])

    # Category 3: Important Labels
    labels_layers = ['FEATURE_LABEL', 'KILOMETRAJE', 'DEFPOINTS']
    labels_df = df[df['Layer'].isin(labels_layers) & df['Text'].notna()].copy()
    labels_df = labels_df[['X', 'Y', 'Z', 'Layer', 'Text', 'SubClasses', 'EntityHandle']].sort_values(by=['Layer', 'Text'])

    # Category 4: Fiber Optic and Project Dimensions (Cotas)
    cotas_layers = ['F.O. COTAS', 'F.O. RETENIDAS', 'COTAS PROYECTO']
    cotas_df = df[df['Layer'].isin(cotas_layers) & df['Text'].notna()].copy()
    cotas_df = cotas_df[['X', 'Y', 'Z', 'Layer', 'Text', 'SubClasses', 'EntityHandle']].sort_values(by=['Layer', 'X'])

    # Category 5: Lines & Polylines
    lines_layers = ['EJE DE VIA', 'CARRETERA', 'DERECHO DE VIA', 'F.O. CABLE']
    lines_df = df[df['Layer'].isin(lines_layers) | df['SubClasses'].str.contains('Polyline|Line', na=False)].copy()
    lines_df = lines_df[['X', 'Y', 'Z', 'Layer', 'SubClasses', 'EntityHandle']].sort_values(by=['Layer', 'X'])

    print("Escribiendo archivo Excel...")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Resumen', index=False)
        poles_df.to_excel(writer, sheet_name='Postes y Puntos', index=False)
        labels_df.to_excel(writer, sheet_name='Etiquetas y Nombres', index=False)
        cotas_df.to_excel(writer, sheet_name='Cotas Fibra Optica', index=False)
        lines_df.to_excel(writer, sheet_name='Lineas y Polilineas', index=False)

    print("Aplicando estilos profesionales al Excel...")
    wb = openpyxl.load_workbook(excel_path)
    
    # Palette colors (Modern Blue theme)
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        
        # Format Headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        ws.row_dimensions[1].height = 25

        # Format Data Rows
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row in range(2, max_row + 1):
            ws.row_dimensions[row].height = 18
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                
                # Format coordinates (X, Y, Z) and text alignments
                col_name = ws.cell(row=1, column=col).value
                if col_name in ['X', 'Y']:
                    cell.number_format = '#,##0.000'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_name == 'Z':
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_name in ['Cantidad de Elementos']:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif col_name in ['Capa (Layer)', 'Layer', 'SubClasses', 'EntityHandle']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                # If coordinate formatted, estimate length
                if cell.row > 1 and ws.cell(row=1, column=cell.column).value in ['X', 'Y', 'Z']:
                    try:
                        val = f"{float(cell.value):,.3f}"
                    except:
                        pass
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    print(f"¡Proceso completado exitosamente! El archivo se guardó en: {excel_path}")

if __name__ == "__main__":
    main()
